"""원본 '통합 문서1.xlsx'는 과거 99개 파일 통합 작업의 잔재(#REF! 정의된 이름 5천여개 등)로
openpyxl로 재저장하면 Excel이 항상 '복구 모드'로만 열리고, 그 과정에서 이미지가 삭제된다.
그래서 원본을 그대로 쓰는 대신 동일한 레이아웃/서식을 가진 깨끗한 템플릿을 새로 생성한다.
이 스크립트는 1회성으로 template.xlsx를 만든다.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

MEDIUM = Side(style="medium")
THIN = Side(style="thin")

COL_WIDTHS = {"A": 2.880802, "J": 2.127564, "K": 2.880802, "X": 2.127564, "Y": 2.880802, "AG": 3.131881}
DEFAULT_COL_WIDTH = 3.131881


def set_border(ws, row, col, left=None, right=None, top=None, bottom=None):
    cell = ws.cell(row=row, column=col)
    b = cell.border
    cell.border = Border(
        left=left if left is not None else b.left,
        right=right if right is not None else b.right,
        top=top if top is not None else b.top,
        bottom=bottom if bottom is not None else b.bottom,
    )


def draw_box(ws, r1, r2, c1, c2, side=MEDIUM, bottom_side=None):
    """bottom_side를 따로 지정하면 사진 박스 아래쪽(캡션과 맞닿는 변)만 다른 굵기로 그릴 수 있다."""
    if bottom_side is None:
        bottom_side = side
    for col in range(c1, c2 + 1):
        set_border(ws, r1, col, top=side)
        set_border(ws, r2, col, bottom=bottom_side)
    for row in range(r1, r2 + 1):
        set_border(ws, row, c1, left=side)
        set_border(ws, row, c2, right=side)


def add_caption_rows(ws, row1, row2, col_start, col_end, label_dwg, label_location, label_content, label_date):
    """사진 한 장 전용 캡션 2행을 만든다: row1=기기번호+작업구역, row2=작업내용+날짜.
    col_start~col_end 범위를 라벨(3칸)+값+라벨(3칸)+값으로 나눈다 (값 두 칸은 남는 폭을 균등 분배)."""
    total = col_end - col_start + 1
    label_w = 3
    value_w = (total - 2 * label_w) // 2
    l1s, l1e = col_start, col_start + label_w - 1
    v1s, v1e = l1e + 1, l1e + value_w
    l2s, l2e = v1e + 1, v1e + label_w
    v2s, v2e = l2e + 1, col_end

    for r in (row1, row2):
        ws.merge_cells(start_row=r, start_column=l1s, end_row=r, end_column=l1e)
        ws.merge_cells(start_row=r, start_column=v1s, end_row=r, end_column=v1e)
        ws.merge_cells(start_row=r, start_column=l2s, end_row=r, end_column=l2e)
        ws.merge_cells(start_row=r, start_column=v2s, end_row=r, end_column=v2e)

    ws.cell(row=row1, column=l1s, value=label_dwg)
    ws.cell(row=row1, column=l2s, value=label_location)
    ws.cell(row=row2, column=l1s, value=label_content)
    ws.cell(row=row2, column=l2s, value=label_date)

    font = Font(name="맑은 고딕", size=10, bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for r in (row1, row2):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.alignment = align

    # 바깥 테두리(좌/우)
    for r in (row1, row2):
        set_border(ws, r, col_start, left=MEDIUM)
        set_border(ws, r, col_end, right=MEDIUM)
    # 사진 바로 아래(내용 시작 지점)는 가는선, 캡션 블록의 맨 아래는 굵은선
    for c in range(col_start, col_end + 1):
        set_border(ws, row1, c, top=THIN)
        set_border(ws, row2, c, bottom=MEDIUM)
    # 행1/행2 사이 구분선(가는선)
    for c in range(col_start, col_end + 1):
        set_border(ws, row1, c, bottom=THIN)
    # 값1<->라벨2 구분선(굵은선), 라벨1<->값1 / 라벨2<->값2 구분선(가는선)
    for r in (row1, row2):
        set_border(ws, r, v1e, right=MEDIUM)
        set_border(ws, r, l2s, left=MEDIUM)
        set_border(ws, r, v1s, left=THIN)
        set_border(ws, r, v2s, left=THIN)

    ws.row_dimensions[row1].height = 18.75
    ws.row_dimensions[row2].height = 18.75

    # 값 칸만 잠금 해제 — 시트 보호가 걸려도 캡션 값(기기번호/작업구역/작업내용/날짜)은 계속 수정 가능해야 한다.
    for r in (row1, row2):
        for c in list(range(v1s, v1e + 1)) + list(range(v2s, v2e + 1)):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    return {
        "dwg": f"{get_column_letter(v1s)}{row1}",
        "location": f"{get_column_letter(v2s)}{row1}",
        "content": f"{get_column_letter(v1s)}{row2}",
        "date": f"{get_column_letter(v2s)}{row2}",
    }


def build_sheet(wb, title, split, labels, margin_lr_in=0.7):
    ws = wb.create_sheet(title)

    ws.sheet_format.defaultColWidth = DEFAULT_COL_WIDTH
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.sheet_format.defaultRowHeight = 15.0

    ws.merge_cells("A1:AF1")
    ws["A1"] = "사 진 대 지 (CONDUIT PIPE)"
    ws["A1"].font = Font(name="HY헤드라인M", size=22)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 37.5

    ws["A2"] = "공사명 : 샤힌 에틸렌시설 건설공사 PKG1 계장공사 (3공구)"
    ws["A2"].font = Font(name="맑은 고딕", size=10, bold=True)
    ws.row_dimensions[2].height = 18.75

    # 페이지 상단 작업내용 요약줄 (해당 페이지 사진들의 작업내용을 모아서 보여줌)
    ws.merge_cells("A3:AF3")
    ws["A3"] = "작업내용 : "
    ws["A3"].font = Font(name="맑은 고딕", size=10, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18.75

    if split:
        # 사진 4장: 사진마다 독립된 캡션(기기번호+작업구역 / 작업내용+날짜, 2행)을 갖는다.
        draw_box(ws, 4, 24, 1, 16, bottom_side=THIN)
        draw_box(ws, 4, 24, 17, 32, bottom_side=THIN)
        draw_box(ws, 27, 47, 1, 16, bottom_side=THIN)
        draw_box(ws, 27, 47, 17, 32, bottom_side=THIN)
        cell_map = {
            "top_left": add_caption_rows(ws, 25, 26, 1, 16, labels["dwg"], labels["location"], labels["content"], labels["date"]),
            "top_right": add_caption_rows(ws, 25, 26, 17, 32, labels["dwg"], labels["location"], labels["content"], labels["date"]),
            "bottom_left": add_caption_rows(ws, 48, 49, 1, 16, labels["dwg"], labels["location"], labels["content"], labels["date"]),
            "bottom_right": add_caption_rows(ws, 48, 49, 17, 32, labels["dwg"], labels["location"], labels["content"], labels["date"]),
        }
    else:
        draw_box(ws, 4, 24, 1, 32, bottom_side=THIN)
        draw_box(ws, 27, 47, 1, 32, bottom_side=THIN)
        cell_map = {
            "top": add_caption_rows(ws, 25, 26, 1, 32, labels["dwg"], labels["location"], labels["content"], labels["date"]),
            "bottom": add_caption_rows(ws, 48, 49, 1, 32, labels["dwg"], labels["location"], labels["content"], labels["date"]),
        }
    print_area_end_row = 49

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:AF{print_area_end_row}"
    ws.page_margins.left = margin_lr_in
    ws.page_margins.right = margin_lr_in
    ws.page_margins.top = 0.4724409448818898
    ws.page_margins.bottom = 0.4724409448818898
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True

    return ws, cell_map


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # 작업사진관리(work-shoot) 시스템과 동일한 필드명 사용, 4장/2장 레이아웃 공통
    common_labels = {"dwg": "기기번호", "location": "작업구역", "content": "작업내용", "date": "날짜"}

    # 사진 칸을 세로 111mm x 가로 94.3mm에 맞추기 위해 좌우 여백을 10.7mm로 줄였다.
    margin_lr_in = 10.7 / 25.4  # 0.4213in

    _, cell_map_4 = build_sheet(wb, "사진4장 (2)", split=True, labels=common_labels, margin_lr_in=margin_lr_in)
    _, cell_map_2 = build_sheet(wb, "사진2장", split=False, labels=common_labels, margin_lr_in=margin_lr_in)

    wb.save("template.xlsx")
    print("template.xlsx generated")
    print("사진4장 cell_map:", cell_map_4)
    print("사진2장 cell_map:", cell_map_2)


if __name__ == "__main__":
    main()
