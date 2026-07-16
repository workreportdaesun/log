"""template.xlsx를 기준으로 사진대지 엑셀을 생성하는 핵심 로직.

원본 템플릿의 두 시트를 그대로 복제해서 쓰기 때문에(서식/병합/테두리/폰트 재구현 없음)
결과물이 원본과 동일한 폼(form)을 유지한다.
"""
import io
import os
import re
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from build_template import COL_WIDTHS, DEFAULT_COL_WIDTH

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "template.xlsx")

SHEET_4 = "사진4장 (2)"  # 사진 4장(2x2) 레이아웃 - 사진마다 독립된 캡션
SHEET_2 = "사진2장"       # 사진 2장(세로) 레이아웃 - 사진마다 독립된 캡션

# 레이아웃별 "세트" 순서 (사진 1장 = 세트 1개, 각자 독립된 dwg/location/content/date를 가짐)
SET_NAMES = {
    "4": ["top_left", "top_right", "bottom_left", "bottom_right"],
    "2": ["top", "bottom"],
}

PRINT_AREA_END_ROW = 49  # 두 시트 모두 행 구조가 동일 (제목/공사명/작업내용요약 + 박스*2 + 캡션*2)

# 캡션 값 셀 주소 (병합 범위의 좌상단 셀). build_template.py의 cell_map과 동일해야 한다.
FIELD_CELLS = {
    "4": {
        "top_left": {"dwg": "D25", "location": "L25", "content": "D26", "date": "L26"},
        "top_right": {"dwg": "T25", "location": "AB25", "content": "T26", "date": "AB26"},
        "bottom_left": {"dwg": "D48", "location": "L48", "content": "D49", "date": "L49"},
        "bottom_right": {"dwg": "T48", "location": "AB48", "content": "T49", "date": "AB49"},
    },
    "2": {
        "top": {"dwg": "D25", "location": "T25", "content": "D26", "date": "T26"},
        "bottom": {"dwg": "D48", "location": "T48", "content": "D49", "date": "T49"},
    },
}
TITLE_CELL = "A1"
PROJECT_CELL = "A2"
SUMMARY_CELL = "A3"  # 페이지 상단 "작업내용 : ..." 요약줄

# build_template.py에서 실제로 지정한 행 높이(pt). 지정 없는 행은 defaultRowHeight(15.0) 사용.
# 사진4장/사진2장 시트가 이제 행 구조가 동일해서 표 하나만 쓴다.
ROW_HEIGHTS_PT = {1: 37.5, 2: 18.75, 3: 18.75, 25: 18.75, 26: 18.75, 48: 18.75, 49: 18.75}
DEFAULT_ROW_HEIGHT_PT = 15.0

EMU_PER_PT = 12700
MDW_PX = 7  # Calibri 11 기준 max digit width
MM_TO_EMU = 36000

# 목표: 사진과 박스 테두리 사이 간격 1mm. 그런데 Excel COM의 PDF 변환 과정에서 TwoCellAnchor
# 이미지가 셀 경계 대비 근소하게 아래/오른쪽으로 밀려 렌더링되는 현상이 실측 결과 확인되어
# (템플릿 실제 인쇄물을 픽셀 단위로 측정: 상/좌측 간격은 명목값보다 작게, 하/우측은 크게 나옴),
# 최종 인쇄 결과가 정확히 1mm가 되도록 변 방향별로 보정한 값을 사용한다.
INSET_TOP_EMU = 46150
INSET_LEFT_EMU = 45240
INSET_RIGHT_EMU = 2760
INSET_BOTTOM_EMU = 10490

# 사진 박스 영역 (1-indexed, build_template.py의 draw_box 호출과 동일)
PHOTO_BOXES_4 = {
    "top_left": (4, 24, 1, 16),
    "top_right": (4, 24, 17, 32),
    "bottom_left": (27, 47, 1, 16),
    "bottom_right": (27, 47, 17, 32),
}
PHOTO_BOXES_2 = {
    "top": (4, 24, 1, 32),
    "bottom": (27, 47, 1, 32),
}


def _col_width_to_emu(width):
    px = int(((256 * width + int(128 / MDW_PX)) / 256) * MDW_PX)
    return px * 9525  # 96 DPI: 1px = 9525 EMU


def _col_offsets_emu(n_cols=33):
    """offsets[c] = 컬럼 c(1-indexed)의 왼쪽 경계 EMU. offsets[0]=0."""
    offsets = [0]
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        width = COL_WIDTHS.get(letter, DEFAULT_COL_WIDTH)
        offsets.append(offsets[-1] + _col_width_to_emu(width))
    return offsets


def _row_offsets_emu(row_heights_pt, n_rows):
    """offsets[r] = 행 r(1-indexed)의 위쪽 경계 EMU. offsets[0]=0."""
    offsets = [0]
    for r in range(1, n_rows + 1):
        h = row_heights_pt.get(r, DEFAULT_ROW_HEIGHT_PT)
        offsets.append(offsets[-1] + h * EMU_PER_PT)
    return offsets


_COL_OFFSETS = _col_offsets_emu()
_ROW_OFFSETS = _row_offsets_emu(ROW_HEIGHTS_PT, PRINT_AREA_END_ROW)


def _emu_to_marker(offsets, absolute_emu):
    """절대 EMU 위치를 (0-indexed index, 그 칸 내부 offset)으로 변환."""
    for i in range(len(offsets) - 1):
        if offsets[i] <= absolute_emu <= offsets[i + 1]:
            return i, int(round(absolute_emu - offsets[i]))
    return len(offsets) - 2, int(round(absolute_emu - offsets[-2]))


def _photo_anchor_box(row_start, row_end, col_start, col_end):
    """box: 1-indexed (row_start, row_end, col_start, col_end) 셀 테두리에서 변별 보정 inset만큼 안쪽으로 들어간 anchor 좌표."""
    left = _COL_OFFSETS[col_start - 1] + INSET_LEFT_EMU
    right = _COL_OFFSETS[col_end] - INSET_RIGHT_EMU
    top = _ROW_OFFSETS[row_start - 1] + INSET_TOP_EMU
    bottom = _ROW_OFFSETS[row_end] - INSET_BOTTOM_EMU

    fc, fco = _emu_to_marker(_COL_OFFSETS, left)
    tc, tco = _emu_to_marker(_COL_OFFSETS, right)
    fr, fro = _emu_to_marker(_ROW_OFFSETS, top)
    tr, tro = _emu_to_marker(_ROW_OFFSETS, bottom)
    return ((fc, fco, fr, fro), (tc, tco, tr, tro))


ANCHORS_4 = {name: _photo_anchor_box(*box) for name, box in PHOTO_BOXES_4.items()}
ANCHORS_2 = {name: _photo_anchor_box(*box) for name, box in PHOTO_BOXES_2.items()}
ANCHORS = {"4": ANCHORS_4, "2": ANCHORS_2}


def _make_anchor(box):
    (fc, fco, fr, fro), (tc, tco, tr, tro) = box
    _from = AnchorMarker(col=fc, colOff=fco, row=fr, rowOff=fro)
    _to = AnchorMarker(col=tc, colOff=tco, row=tr, rowOff=tro)
    return TwoCellAnchor(editAs="oneCell", _from=_from, to=_to)


def _shift_anchor_box(box, row_offset):
    """페이지 블록마다 행 구조(ROW_HEIGHTS_PT)가 동일하게 반복되므로, 앵커의 행 인덱스만
    블록 크기(PRINT_AREA_END_ROW)만큼 밀면 된다 — 칸 안에서의 세부 오프셋(rowOff)은 그대로 재사용 가능."""
    (fc, fco, fr, fro), (tc, tco, tr, tro) = box
    return ((fc, fco, fr + row_offset, fro), (tc, tco, tr + row_offset, tro))


_CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")


def _shift_cell(addr, row_offset):
    col, row = _CELL_RE.match(addr).groups()
    return f"{col}{int(row) + row_offset}"


def _copy_template_block(src_ws, dest_ws, row_offset, max_col=32):
    """src_ws(사진4장/사진2장 템플릿 시트)의 1~PRINT_AREA_END_ROW행 서식·값·병합을
    dest_ws의 (row_offset+1)~(row_offset+PRINT_AREA_END_ROW)행에 그대로 복제한다."""
    for row in range(1, PRINT_AREA_END_ROW + 1):
        src_dim = src_ws.row_dimensions.get(row)
        if src_dim and src_dim.height is not None:
            dest_ws.row_dimensions[row + row_offset].height = src_dim.height
        for col in range(1, max_col + 1):
            src_cell = src_ws.cell(row=row, column=col)
            dest_cell = dest_ws.cell(row=row + row_offset, column=col)
            dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.number_format = src_cell.number_format
                dest_cell.protection = copy(src_cell.protection)

    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.max_row <= PRINT_AREA_END_ROW:
            dest_ws.merge_cells(
                start_row=merged_range.min_row + row_offset,
                start_column=merged_range.min_col,
                end_row=merged_range.max_row + row_offset,
                end_column=merged_range.max_col,
            )


def _add_image(ws, image_bytes, anchor_box):
    img = XLImage(io.BytesIO(image_bytes))
    img.anchor = _make_anchor(anchor_box)
    ws.add_image(img)


def _set_fields(ws, cells, fields):
    """fields: dict with keys dwg, location, content, date (모두 optional)"""
    if fields.get("dwg"):
        ws[cells["dwg"]] = fields["dwg"]
    if fields.get("location"):
        ws[cells["location"]] = fields["location"]
    if fields.get("content"):
        ws[cells["content"]] = fields["content"]
    if fields.get("date"):
        ws[cells["date"]] = fields["date"]


def _protect_format(ws):
    """서식(행 높이/열 너비/행·열 삽입삭제/사진 위치)만 잠그고 캡션 값은 계속 수정 가능하게 둔다.
    비밀번호는 걸지 않음 — 실수로 서식이 바뀌는 걸 막는 용도이지 편집을 완전히 막는 용도가 아니다.
    캡션 값 셀은 build_template.py의 add_caption_rows에서 이미 Protection(locked=False)로 표시해뒀다."""
    ws.protection.sheet = True
    ws.protection.formatCells = True
    ws.protection.formatColumns = True
    ws.protection.formatRows = True
    ws.protection.insertColumns = True
    ws.protection.insertRows = True
    ws.protection.deleteColumns = True
    ws.protection.deleteRows = True
    ws.protection.objects = True  # 사진(도형) 이동/크기조절도 금지


def _content_summary(sets):
    """페이지 안 사진들의 작업내용을 중복 없이 모아 '작업내용 : A, B' 형태로 만든다."""
    seen = []
    for s in sets or []:
        content = (s or {}).get("content", "").strip()
        if content and content not in seen:
            seen.append(content)
    if not seen:
        return None
    return "작업내용 : " + ", ".join(seen)


def build_workbook(pages, settings):
    """pages: list of dict, each:
        {
          "layout": "4" | "2",
          "sets": [
            {"dwg":.., "location":.., "content":.., "date":.., "photo": bytes|None},
            ...  (레이아웃 4는 4개: top_left, top_right, bottom_left, bottom_right 순서
                   레이아웃 2는 2개: top, bottom 순서)
          ],
        }
    settings: {"project_name": str, "company_name": str, "work_title": str}
    반환: openpyxl.Workbook (모든 페이지가 시트 하나에 순서대로 쌓인 상태. 원본 템플릿 시트는 삭제됨)
    """
    wb = load_workbook(TEMPLATE_PATH)
    src_sheets = {SHEET_4: wb[SHEET_4], SHEET_2: wb[SHEET_2]}

    title = f"사 진 대 지 ({settings.get('work_title', '').strip()})" if settings.get("work_title") else None
    project_bits = []
    if settings.get("project_name"):
        project_bits.append(f"공사명 : {settings['project_name'].strip()}")
    if settings.get("company_name"):
        project_bits.append(f"회사명 : {settings['company_name'].strip()}")
    project_line = "    ".join(project_bits) if project_bits else None

    dest_ws = wb.create_sheet("사진대지")
    dest_ws.sheet_format.defaultColWidth = DEFAULT_COL_WIDTH
    for col, w in COL_WIDTHS.items():
        dest_ws.column_dimensions[col].width = w
    dest_ws.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT_PT

    for idx, page in enumerate(pages):
        row_offset = idx * PRINT_AREA_END_ROW
        layout = page.get("layout", "4")
        src_ws = src_sheets[SHEET_4 if layout == "4" else SHEET_2]
        _copy_template_block(src_ws, dest_ws, row_offset)

        if title:
            dest_ws[_shift_cell(TITLE_CELL, row_offset)] = title
        if project_line:
            dest_ws[_shift_cell(PROJECT_CELL, row_offset)] = project_line

        sets = page.get("sets") or []
        summary = _content_summary(sets)
        if summary:
            dest_ws[_shift_cell(SUMMARY_CELL, row_offset)] = summary

        for set_name, set_data in zip(SET_NAMES[layout], sets):
            set_data = set_data or {}
            shifted_cells = {k: _shift_cell(v, row_offset) for k, v in FIELD_CELLS[layout][set_name].items()}
            _set_fields(dest_ws, shifted_cells, set_data)
            photo = set_data.get("photo")
            if photo:
                _add_image(dest_ws, photo, _shift_anchor_box(ANCHORS[layout][set_name], row_offset))

        # 마지막 페이지가 아니면 이 블록 끝에 페이지 나누기를 넣어서, 인쇄/PDF 변환 시
        # 한 시트 안에서도 페이지별로 계속 나뉘어 출력되게 한다.
        if idx < len(pages) - 1:
            dest_ws.row_breaks.append(Break(id=row_offset + PRINT_AREA_END_ROW))

    total_rows = max(len(pages), 1) * PRINT_AREA_END_ROW
    dest_ws.print_area = f"A1:AF{total_rows}"
    dest_ws.page_setup.orientation = "portrait"
    dest_ws.page_setup.paperSize = 9
    dest_ws.page_setup.fitToWidth = 1
    dest_ws.page_setup.fitToHeight = 0  # 세로는 제한 없이 페이지 나누기 기준으로 여러 장에 걸쳐 출력
    dest_ws.sheet_properties.pageSetUpPr.fitToPage = True
    _protect_format(dest_ws)

    # 원본 템플릿 시트 제거 (완성된 시트 하나만 남김)
    del wb[SHEET_4]
    del wb[SHEET_2]

    return wb
